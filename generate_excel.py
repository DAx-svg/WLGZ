"""
来料出入库管理系统 - Excel 生成脚本
====================================
使用 openpyxl 生成 WPS Office 兼容的出入库管理 Excel 文件。
纯公式驱动，无需 VBA 宏。

用法：
  python generate_excel.py          生成带示例数据的模板
  python generate_excel.py --clear  生成空白模板（清除示例数据）
输出：来料出入库管理系统.xlsx
"""

import os
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers, Protection
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.workbook.defined_name import DefinedName

# ============================================================================
# 常量 & 样式
# ============================================================================

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           '来料出入库管理系统.xlsx')

# 颜色
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color='1F4E79')
KPI_LABEL_FONT = Font(name='微软雅黑', size=9, color='666666')
KPI_VALUE_FONT = Font(name='微软雅黑', size=22, bold=True, color='1F4E79')

ODD_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
EVEN_FILL = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
DANGER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GOOD_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
DASHBOARD_CARD_FILL = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
SECTION_FILL = PatternFill(start_color='E9EDF4', end_color='E9EDF4', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4'),
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='2F5496'),
    right=Side(style='thin', color='2F5496'),
    top=Side(style='thin', color='2F5496'),
    bottom=Side(style='thin', color='2F5496'),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_CENTER = Alignment(horizontal='right', vertical='center')


def style_header_row(ws, num_cols, row=1):
    """给表头行应用样式"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER


def style_data_rows(ws, start_row, end_row, num_cols):
    """给数据行应用交替颜色和边框"""
    for r in range(start_row, end_row + 1):
        fill = EVEN_FILL if (r - start_row) % 2 == 1 else ODD_FILL
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if c > 1:
                cell.alignment = LEFT_WRAP
            else:
                cell.alignment = CENTER


def apply_borders(ws, min_row, max_row, min_col, max_col):
    """给区域加边框"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def set_col_widths(ws, widths):
    """按列设置宽度，widths: {letter: width}"""
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


# ============================================================================
# Sheet 1: 操作说明
# ============================================================================

def build_instructions(wb):
    ws = wb.active
    ws.title = '操作说明'

    set_col_widths(ws, {'A': 20, 'B': 50, 'C': 20})

    # 标题
    ws.merge_cells('A1:C1')
    title = ws['A1']
    title.value = '📦 来料出入库管理系统 — 操作说明'
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    sections = [
        ('一、系统概述', [
            '本系统是一套基于 Excel 公式的仓库出入库管理工具，适用于物料种类在 200 种以内的中小型仓库。',
            '所有库存数据由公式自动计算，无需手动汇总，有效避免人为错误。',
            '兼容 WPS Office 2019+ 和 Microsoft Excel 2016+，无需安装任何插件。',
        ]),
        ('二、使用流程', [
            '① 在「物料主数据」表中录入所有需要管理的物料信息（编码、名称、规格、安全库存等）。',
            '② 物料到货时，在「入库记录」表中登记（日期、物料编码、数量、供应商等）。',
            '③ 物料领用/发出时，在「出库记录」表中登记（日期、物料编码、数量、领用部门等）。',
            '④ 在「库存汇总」表中查看各物料的实时库存和状态预警。',
            '⑤ 在「仪表盘」中查看整体 KPI 概览和月度趋势。',
        ]),
        ('三、各工作表说明', [
            '【物料主数据】所有物料的基础信息。编码必须唯一，后续入库/出库通过编码引用。D列(单位)和E列(类别)有下拉菜单。',
            '【入库记录】记录每次物料的入库明细。C列通过下拉菜单选择物料编码，D列名称自动填充。E列数量须为正整数。',
            '【出库记录】记录每次物料的出库明细。操作方式与入库记录相同。',
            '【库存汇总】根据入库和出库记录自动计算每种物料的当前库存。H列显示实时库存，I列显示库存状态（正常/低于安全库存/缺货）。',
            '【仪表盘】展示关键指标卡、低库存预警列表和月度出入库趋势图，一目了然。',
        ]),
        ('四、注意事项', [
            '⚠️ 物料编码必须唯一，不要重复。新增物料时请先检查「物料主数据」表中是否已存在。',
            '⚠️ 入库和出库记录中，物料名称列（D列）由 VLOOKUP 公式自动填充，请勿手动修改。',
            '⚠️ 日期请使用 YYYY-MM-DD 格式（如 2026-06-15），避免使用中文日期格式。',
            '⚠️ 如需添加更多数据行，请在数据区域中间「插入行」（不是在末尾追加），以确保公式范围覆盖新行。',
            '⚠️「库存汇总」和「仪表盘」中的数据均为公式自动计算，请勿手动修改。',
            '⚠️ 关闭文件时请保存，WPS 会提示是否保存更改，点击"是"即可。',
        ]),
        ('五、常见问题', [
            'Q: 入库时输入物料编码后，名称没有自动出现？',
            'A: 请检查该编码在「物料主数据」表中是否存在，编码必须完全一致（包括大小写和连字符）。',
            '',
            'Q: 库存汇总中显示"缺货"但实际有库存？',
            'A: 请检查是否所有入库和出库记录都已正确录入。库存 = 累计入库 - 累计出库。',
            '',
            'Q: 如何备份数据？',
            'A: 建议定期将文件另存为带日期后缀的副本，如"来料出入库管理系统_20260615.xlsx"。',
            '',
            'Q: 可以多人同时编辑吗？',
            'A: Excel 文件不支持多人实时协作。建议一人编辑，其他人以只读方式查看。',
        ]),
    ]

    row = 3
    for title_text, lines in sections:
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws.cell(row=row, column=1, value=title_text)
        cell.font = Font(name='微软雅黑', size=13, bold=True, color='1F4E79')
        ws.row_dimensions[row].height = 28
        row += 1

        for line in lines:
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws.cell(row=row, column=1, value=line)
            if line.startswith('Q:'):
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
            elif line.startswith('A:'):
                cell.font = Font(name='微软雅黑', size=10, color='333333')
            elif line.startswith('⚠️'):
                cell.font = Font(name='微软雅黑', size=10, color='BF8F00')
            elif line.startswith('【'):
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='2F5496')
            elif line and line[0] in '①②③④⑤':
                cell.font = Font(name='微软雅黑', size=10, color='333333')
            else:
                cell.font = Font(name='微软雅黑', size=10, color='555555')
            cell.alignment = LEFT_WRAP
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1  # 段落间距

    # 页脚
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1,
                   value=f'文档生成日期：{date.today()}  |  版本 1.0  |  适用于 WPS Office / Excel')
    cell.font = Font(name='微软雅黑', size=9, italic=True, color='999999')
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.sheet_properties.tabColor = '4472C4'
    return ws


# ============================================================================
# Sheet 2: 物料主数据
# ============================================================================

# 示例物料数据
SAMPLE_MATERIALS = [
    ('MAT-001', '螺栓 M8×30', 'GB/T 5782-2000', '个', '原材料', 200, '常用紧固件'),
    ('MAT-002', '螺母 M8', 'GB/T 6170-2000', '个', '原材料', 300, '配套MAT-001'),
    ('MAT-003', '弹簧垫圈 M8', 'GB/T 93-1987', '个', '原材料', 500, ''),
    ('MAT-004', '钢板 Q235', '2.0mm×1250×2500', 'kg', '原材料', 500, '定制尺寸'),
    ('MAT-005', 'PCB控制板 V2.1', 'FR-4 双面板', '台', '半成品', 20, '核心控制模块'),
    ('MAT-006', '电源模块 24V/5A', 'LRS-150-24', '台', '半成品', 15, '明纬电源'),
    ('MAT-007', '包装纸箱', '400×300×200mm', '个', '包装材料', 100, '三层瓦楞'),
    ('MAT-008', '成品设备 Model-X', 'X-2000标准型', '台', '成品', 10, '整机成品'),
]

UNITS = '个,箱,套,kg,米,卷,包,桶,台,件'
CATEGORIES = '原材料,半成品,成品,包装材料,辅料,备品备件'

MATERIAL_HEADERS = ['物料编码', '物料名称', '规格型号', '单位', '物料类别', '安全库存', '备注']
MATERIAL_COLS = ['A', 'B', 'C', 'D', 'E', 'F', 'G']


def build_material_master(wb, sample_data=True):
    ws = wb.create_sheet('物料主数据')

    # 列宽
    set_col_widths(ws, {'A': 14, 'B': 22, 'C': 24, 'D': 10, 'E': 14, 'F': 12, 'G': 22})

    # 表头
    for i, h in enumerate(MATERIAL_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, len(MATERIAL_HEADERS))
    ws.row_dimensions[1].height = 28

    # 示例数据（仅在非清除模式下写入）
    if sample_data:
        for r, mat in enumerate(SAMPLE_MATERIALS, 2):
            for c, val in enumerate(mat, 1):
                ws.cell(row=r, column=c, value=val)

    # 预格式化 200 行
    NUM_ROWS = 200
    style_data_rows(ws, 2, NUM_ROWS + 1, len(MATERIAL_HEADERS))

    # 居中显示的列
    for r in range(2, NUM_ROWS + 2):
        for c in [1, 4, 5, 6]:  # 编码、单位、类别、安全库存居中
            ws.cell(row=r, column=c).alignment = CENTER

    # 冻结窗格
    ws.freeze_panes = 'A2'

    # 数据验证：单位（D列）
    dv_units = DataValidation(type='list', formula1=f'"{UNITS}"', allow_blank=True)
    dv_units.error = '请选择有效的单位'
    dv_units.errorTitle = '输入错误'
    ws.add_data_validation(dv_units)
    dv_units.add(f'D2:D{NUM_ROWS + 1}')

    # 数据验证：类别（E列）
    dv_cats = DataValidation(type='list', formula1=f'"{CATEGORIES}"', allow_blank=True)
    dv_cats.error = '请选择有效的物料类别'
    dv_cats.errorTitle = '输入错误'
    ws.add_data_validation(dv_cats)
    dv_cats.add(f'E2:E{NUM_ROWS + 1}')

    # 数据验证：安全库存（F列）正整数
    dv_safety = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='0')
    dv_safety.error = '安全库存必须为大于等于0的整数'
    ws.add_data_validation(dv_safety)
    dv_safety.add(f'F2:F{NUM_ROWS + 1}')

    # 重复编码检测（隐藏列H）
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=8,
                value=f'=IF(A{r}="","",IF(COUNTIF($A$2:$A${NUM_ROWS + 1},A{r})>1,"重复编码!",""))')
        ws.cell(row=r, column=8).font = Font(name='微软雅黑', size=9, color='FF0000')
    ws.column_dimensions['H'].hidden = True

    # 定义命名区域 MaterialCodes
    ref = f'物料主数据!$A$2:$A${NUM_ROWS + 1}'
    dn = DefinedName('MaterialCodes', attr_text=ref)
    wb.defined_names.add(dn)

    # 页脚提示
    footer_row = NUM_ROWS + 3
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    cell = ws.cell(row=footer_row, column=1,
                   value='💡 提示：物料编码请保持唯一；D列(单位)和E列(类别)可通过下拉菜单选择；新增物料在已有数据下方空白行填入即可，库存汇总表会自动同步。')
    cell.font = Font(name='微软雅黑', size=9, italic=True, color='999999')

    ws.sheet_properties.tabColor = '5B9BD5'
    return ws


# ============================================================================
# Sheet 3: 入库记录
# ============================================================================

SAMPLE_INBOUND = [
    ('2025-11-05', 'RK-2025-001', 'MAT-001', 1000, '华东五金供应有限公司', 'A-01-03', '张伟', '首批入库'),
    ('2025-11-05', 'RK-2025-002', 'MAT-002', 1500, '华东五金供应有限公司', 'A-01-04', '张伟', ''),
    ('2025-11-15', 'RK-2025-003', 'MAT-003', 800, '宝钢贸易有限公司', 'B-02-01', '李强', ''),
    ('2025-12-01', 'RK-2025-004', 'MAT-004', 50, '深圳华强电子有限公司', 'C-01-01', '王芳', 'V2.1版本'),
    ('2025-12-10', 'RK-2025-005', 'MAT-005', 40, '广州明纬电子有限公司', 'C-01-02', '王芳', ''),
    ('2025-12-20', 'RK-2025-006', 'MAT-006', 300, '杭州包装制品厂', 'D-01-01', '赵军', ''),
    ('2026-01-05', 'RK-2026-001', 'MAT-007', 25, '上海自动化设备有限公司', 'E-01-01', '张伟', '已完成质检'),
    ('2026-01-20', 'RK-2026-002', 'MAT-008', 8, '上海自动化设备有限公司', 'E-01-02', '张伟', ''),
    ('2026-02-15', 'RK-2026-003', 'MAT-001', 800, '华东五金供应有限公司', 'A-01-03', '李强', '补货'),
    ('2026-03-10', 'RK-2026-004', 'MAT-003', 600, '宝钢贸易有限公司', 'B-02-01', '李强', ''),
    ('2026-03-22', 'RK-2026-005', 'MAT-005', 30, '广州明纬电子有限公司', 'C-01-02', '王芳', ''),
    ('2026-04-01', 'RK-2026-006', 'MAT-004', 40, '深圳华强电子有限公司', 'C-01-01', '赵军', 'V2.1批次'),
    ('2026-04-18', 'RK-2026-007', 'MAT-002', 1200, '华东五金供应有限公司', 'A-01-04', '张伟', ''),
    ('2026-05-05', 'RK-2026-008', 'MAT-006', 250, '杭州包装制品厂', 'D-01-01', '赵军', ''),
    ('2026-05-20', 'RK-2026-009', 'MAT-007', 15, '上海自动化设备有限公司', 'E-01-01', '李强', '紧急订单'),
]

INBOUND_HEADERS = ['日期', '单据编号', '物料编码', '物料名称',
                   '入库数量', '供应商', '库位', '操作人', '备注']
INBOUND_COLS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']


def build_inbound_records(wb, sample_data=True):
    ws = wb.create_sheet('入库记录')
    NUM_COLS = len(INBOUND_HEADERS)
    DATA_ROWS = 500

    set_col_widths(ws, {'A': 13, 'B': 16, 'C': 14, 'D': 22,
                        'E': 12, 'F': 24, 'G': 12, 'H': 10, 'I': 20})

    # 表头
    for i, h in enumerate(INBOUND_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, NUM_COLS)
    ws.row_dimensions[1].height = 28

    # 示例数据（跳过D列——VLOOKUP公式自动填充物料名称）
    if sample_data:
        for r, rec in enumerate(SAMPLE_INBOUND, 2):
            for i, val in enumerate(rec):
                col = i + 1 if i < 3 else i + 2  # 跳过第4列(D列)
                ws.cell(row=r, column=col, value=val)

    # 预格式化 DATA_ROWS 行
    style_data_rows(ws, 2, DATA_ROWS + 1, NUM_COLS)

    # 居中列
    for r in range(2, DATA_ROWS + 2):
        for c in [1, 3, 5, 7, 8]:
            ws.cell(row=r, column=c).alignment = CENTER

    # D 列 VLOOKUP 公式（自动填充物料名称）
    for r in range(2, DATA_ROWS + 2):
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}="","",IFERROR(VLOOKUP(C{r},物料主数据!A:B,2,FALSE),""))')

    # 冻结窗格
    ws.freeze_panes = 'A2'

    # 数据验证：物料编码（C列）—— 引用命名区域
    dv_code = DataValidation(type='list', formula1='=MaterialCodes', allow_blank=True)
    dv_code.error = '请从下拉列表中选择有效的物料编码，或在物料主数据表中先添加。'
    dv_code.errorTitle = '物料编码无效'
    ws.add_data_validation(dv_code)
    dv_code.add(f'C2:C{DATA_ROWS + 1}')

    # 数据验证：入库数量（E列）正整数
    dv_qty = DataValidation(type='whole', operator='greaterThan', formula1='0')
    dv_qty.error = '入库数量必须为正整数'
    dv_qty.errorTitle = '数量错误'
    ws.add_data_validation(dv_qty)
    dv_qty.add(f'E2:E{DATA_ROWS + 1}')

    # 数据验证：日期（A列）
    dv_date = DataValidation(type='date', operator='greaterThan', formula1='2020-01-01')
    dv_date.error = '请输入有效日期，格式：YYYY-MM-DD'
    dv_date.errorTitle = '日期格式错误'
    ws.add_data_validation(dv_date)
    dv_date.add(f'A2:A{DATA_ROWS + 1}')

    # SUBTOTAL 汇总行
    total_row = DATA_ROWS + 3
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws.cell(row=total_row, column=1, value='入库总计（筛选后可见汇总）').font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = RIGHT_CENTER
    ws.cell(row=total_row, column=5,
            value=f'=SUBTOTAL(9,E2:E{DATA_ROWS + 1})').font = BOLD_FONT
    ws.cell(row=total_row, column=5).alignment = CENTER
    for c in range(1, NUM_COLS + 1):
        ws.cell(row=total_row, column=c).fill = SECTION_FILL

    ws.sheet_properties.tabColor = '70AD47'
    return ws


# ============================================================================
# Sheet 4: 出库记录
# ============================================================================

SAMPLE_OUTBOUND = [
    ('2025-12-05', 'CK-2025-001', 'MAT-001', 300, '生产部-组装组', '刘洋', '生产线日常领用'),
    ('2025-12-08', 'CK-2025-002', 'MAT-002', 400, '生产部-组装组', '刘洋', ''),
    ('2025-12-15', 'CK-2025-003', 'MAT-003', 200, '生产部-冲压组', '陈明', ''),
    ('2026-01-10', 'CK-2026-001', 'MAT-001', 500, '生产部-组装组', '刘洋', ''),
    ('2026-01-25', 'CK-2026-002', 'MAT-004', 25, '研发部', '周杰', '研发测试用'),
    ('2026-02-20', 'CK-2026-003', 'MAT-005', 20, '售后维修部', '吴峰', '售后维修备件'),
    ('2026-03-01', 'CK-2026-004', 'MAT-007', 10, '销售部', '郑丽', '订单发货 SH-20260301'),
    ('2026-03-15', 'CK-2026-005', 'MAT-002', 500, '生产部-组装组', '刘洋', ''),
    ('2026-04-05', 'CK-2026-006', 'MAT-003', 400, '生产部-冲压组', '陈明', ''),
    ('2026-04-20', 'CK-2026-007', 'MAT-005', 25, '售后维修部', '吴峰', ''),
    ('2026-05-10', 'CK-2026-008', 'MAT-006', 200, '包装组', '周建', '发货包装'),
    ('2026-05-25', 'CK-2026-009', 'MAT-001', 600, '生产部-组装组', '刘洋', '紧急生产计划'),
]

OUTBOUND_HEADERS = ['日期', '单据编号', '物料编码', '物料名称',
                    '出库数量', '领用部门/人', '操作人', '备注']
OUTBOUND_COLS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']


def build_outbound_records(wb, sample_data=True):
    ws = wb.create_sheet('出库记录')
    NUM_COLS = len(OUTBOUND_HEADERS)
    DATA_ROWS = 500

    set_col_widths(ws, {'A': 13, 'B': 16, 'C': 14, 'D': 22,
                        'E': 12, 'F': 22, 'G': 10, 'H': 20})

    # 表头
    for i, h in enumerate(OUTBOUND_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, NUM_COLS)
    ws.row_dimensions[1].height = 28

    # 示例数据（跳过D列——VLOOKUP公式自动填充物料名称）
    if sample_data:
        for r, rec in enumerate(SAMPLE_OUTBOUND, 2):
            for i, val in enumerate(rec):
                col = i + 1 if i < 3 else i + 2  # 跳过第4列(D列)
                ws.cell(row=r, column=col, value=val)

    # 预格式化
    style_data_rows(ws, 2, DATA_ROWS + 1, NUM_COLS)

    for r in range(2, DATA_ROWS + 2):
        for c in [1, 3, 5, 7]:
            ws.cell(row=r, column=c).alignment = CENTER

    # D 列 VLOOKUP
    for r in range(2, DATA_ROWS + 2):
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}="","",IFERROR(VLOOKUP(C{r},物料主数据!A:B,2,FALSE),""))')

    ws.freeze_panes = 'A2'

    # 数据验证：物料编码
    dv_code = DataValidation(type='list', formula1='=MaterialCodes', allow_blank=True)
    dv_code.error = '请从下拉列表中选择有效的物料编码。'
    dv_code.errorTitle = '物料编码无效'
    ws.add_data_validation(dv_code)
    dv_code.add(f'C2:C{DATA_ROWS + 1}')

    # 数据验证：出库数量
    dv_qty = DataValidation(type='whole', operator='greaterThan', formula1='0')
    dv_qty.error = '出库数量必须为正整数'
    ws.add_data_validation(dv_qty)
    dv_qty.add(f'E2:E{DATA_ROWS + 1}')

    # 日期验证
    dv_date = DataValidation(type='date', operator='greaterThan', formula1='2020-01-01')
    dv_date.error = '请输入有效日期，格式：YYYY-MM-DD'
    ws.add_data_validation(dv_date)
    dv_date.add(f'A2:A{DATA_ROWS + 1}')

    # SUBTOTAL 汇总
    total_row = DATA_ROWS + 3
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws.cell(row=total_row, column=1, value='出库总计（筛选后可见汇总）').font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = RIGHT_CENTER
    ws.cell(row=total_row, column=5,
            value=f'=SUBTOTAL(9,E2:E{DATA_ROWS + 1})').font = BOLD_FONT
    ws.cell(row=total_row, column=5).alignment = CENTER
    for c in range(1, NUM_COLS + 1):
        ws.cell(row=total_row, column=c).fill = SECTION_FILL

    ws.sheet_properties.tabColor = 'ED7D31'
    return ws


# ============================================================================
# Sheet 5: 库存汇总
# ============================================================================

INV_HEADERS = ['物料编码', '物料名称', '规格型号', '单位', '安全库存',
               '累计入库', '累计出库', '当前库存', '库存状态']
INV_COLS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']


def build_inventory_summary(wb):
    ws = wb.create_sheet('库存汇总')
    NUM_COLS = len(INV_HEADERS)
    NUM_ROWS = 200

    set_col_widths(ws, {'A': 14, 'B': 22, 'C': 24, 'D': 8, 'E': 10,
                        'F': 12, 'G': 12, 'H': 12, 'I': 16})

    # 表头
    for i, h in enumerate(INV_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, NUM_COLS)
    ws.row_dimensions[1].height = 28

    # A 列：引用物料主数据
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=1,
                value=f'=IF(物料主数据!A{r}="","",物料主数据!A{r})')

    # B 列：VLOOKUP 物料名称
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=2,
                value=f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},物料主数据!A:G,2,FALSE),""))')

    # C 列：VLOOKUP 规格型号
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=3,
                value=f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},物料主数据!A:G,3,FALSE),""))')

    # D 列：VLOOKUP 单位
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=4,
                value=f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},物料主数据!A:G,4,FALSE),""))')

    # E 列：VLOOKUP 安全库存
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=5,
                value=f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},物料主数据!A:G,6,FALSE),""))')

    # F 列：累计入库 SUMIF
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=6,
                value=f'=IF(A{r}="","",SUMIF(入库记录!C:C,A{r},入库记录!E:E))')

    # G 列：累计出库 SUMIF
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=7,
                value=f'=IF(A{r}="","",SUMIF(出库记录!C:C,A{r},出库记录!E:E))')

    # H 列：当前库存 = 累计入库 - 累计出库
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=8,
                value=f'=IF(A{r}="","",F{r}-G{r})')

    # I 列：库存状态判断
    for r in range(2, NUM_ROWS + 2):
        ws.cell(row=r, column=9,
                value=f'=IF(A{r}="","",IF(H{r}<=0,"缺货",IF(H{r}<=E{r},"低于安全库存","正常")))')

    # 样式
    style_data_rows(ws, 2, NUM_ROWS + 1, NUM_COLS)

    # 居中列
    for r in range(2, NUM_ROWS + 2):
        for c in [1, 4, 5, 6, 7, 8, 9]:
            ws.cell(row=r, column=c).alignment = CENTER

    # 冻结窗格
    ws.freeze_panes = 'A2'

    # ===== 条件格式 =====

    # 1. 缺货 → 红色
    range_str = f'A2:I{NUM_ROWS + 1}'
    ws.conditional_formatting.add(
        range_str,
        FormulaRule(
            formula=[f'AND($A2<>"", $I2="缺货")'],
            fill=DANGER_FILL,
            font=Font(name='微软雅黑', size=10, color='9C0006', bold=True)
        )
    )

    # 2. 低于安全库存 → 黄色
    ws.conditional_formatting.add(
        range_str,
        FormulaRule(
            formula=[f'AND($A2<>"", $I2="低于安全库存")'],
            fill=WARN_FILL,
            font=Font(name='微软雅黑', size=10, color='9C6500', bold=True)
        )
    )

    # 3. 正常 → 绿色
    ws.conditional_formatting.add(
        range_str,
        FormulaRule(
            formula=[f'AND($A2<>"", $I2="正常")'],
            fill=GOOD_FILL,
            font=Font(name='微软雅黑', size=10, color='006100')
        )
    )

    # 4. 当前库存列（H列）数据条
    ws.conditional_formatting.add(
        f'H2:H{NUM_ROWS + 1}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'))
    )

    # 汇总行
    total_row = NUM_ROWS + 3
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws.cell(row=total_row, column=1, value='📊 汇总统计').font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = RIGHT_CENTER
    ws.cell(row=total_row, column=6,
            value=f'=SUM(F2:F{NUM_ROWS + 1})').font = BOLD_FONT
    ws.cell(row=total_row, column=6).alignment = CENTER
    ws.cell(row=total_row, column=7,
            value=f'=SUM(G2:G{NUM_ROWS + 1})').font = BOLD_FONT
    ws.cell(row=total_row, column=7).alignment = CENTER
    ws.cell(row=total_row, column=8,
            value=f'=SUM(H2:H{NUM_ROWS + 1})').font = BOLD_FONT
    ws.cell(row=total_row, column=8).alignment = CENTER
    for c in range(1, NUM_COLS + 1):
        ws.cell(row=total_row, column=c).fill = SECTION_FILL

    # 页脚提示
    note_row = total_row + 2
    ws.merge_cells(f'A{note_row}:I{note_row}')
    ws.cell(row=note_row, column=1,
            value='💡 提示：本表数据由公式自动生成，无需手动编辑。新增物料后会自动出现在列表中。颜色标识：绿色=正常 黄色=低于安全库存 红色=缺货').font = Font(name='微软雅黑', size=9, italic=True, color='999999')

    ws.sheet_properties.tabColor = 'FFC000'
    return ws


# ============================================================================
# Sheet 6: 仪表盘
# ============================================================================

def build_dashboard(wb):
    ws = wb.create_sheet('仪表盘')
    NUM_INV_ROWS = 200

    set_col_widths(ws, {'A': 3, 'B': 22, 'C': 22, 'D': 22,
                        'E': 22, 'F': 22, 'G': 22, 'H': 22, 'I': 22, 'J': 22,
                        'K': 22, 'L': 22, 'M': 3})

    # ---- 标题 ----
    ws.merge_cells('B1:L1')
    title = ws['B1']
    title.value = '📊 仓库管理仪表盘'
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45

    # ---- KPI 卡片构建辅助函数 ----
    def make_kpi_card(start_col_letter, start_row, end_col_letter, label, formula, accent_color):
        """创建单个 KPI 卡片：值行 + 标签行，合并但不冲突"""
        col_idx = ord(start_col_letter) - ord('A') + 1
        end_col_idx = ord(end_col_letter) - ord('A') + 1
        ncols = end_col_idx - col_idx + 1

        # 先设置背景和边框（在合并之前对每个单元格设置）
        for rr in range(start_row, start_row + 3):
            for cc in range(col_idx, end_col_idx + 1):
                ws.cell(row=rr, column=cc).fill = DASHBOARD_CARD_FILL
                ws.cell(row=rr, column=cc).border = THIN_BORDER

        # 值行合并（第1行）
        val_merge = f'{start_col_letter}{start_row}:{end_col_letter}{start_row}'
        ws.merge_cells(val_merge)
        val_cell = ws.cell(row=start_row, column=col_idx, value=formula)
        val_cell.font = KPI_VALUE_FONT
        val_cell.alignment = Alignment(horizontal='center', vertical='bottom')

        # 标签行合并（第2行）
        lbl_merge = f'{start_col_letter}{start_row + 1}:{end_col_letter}{start_row + 1}'
        ws.merge_cells(lbl_merge)
        lbl_cell = ws.cell(row=start_row + 1, column=col_idx, value=label)
        lbl_cell.font = KPI_LABEL_FONT
        lbl_cell.alignment = Alignment(horizontal='center', vertical='top')

        # 第3行留空（仅背景+边框，已在上面设置）
        ws.row_dimensions[start_row].height = 45
        ws.row_dimensions[start_row + 1].height = 22
        ws.row_dimensions[start_row + 2].height = 8

    # ---- KPI 卡片 (第1行: 3-5) ----
    make_kpi_card('B', 3, 'D', '物料种类总数',
                  f'=SUMPRODUCT((物料主数据!A2:A{NUM_INV_ROWS + 1}<>"")*1)',
                  '4472C4')
    make_kpi_card('F', 3, 'H', '本月入库批次',
                  f'=SUMPRODUCT((入库记录!C2:C501<>"")*(入库记录!E2:E501>0)*(MONTH(入库记录!A2:A501)=MONTH(TODAY()))*(YEAR(入库记录!A2:A501)=YEAR(TODAY())))',
                  '70AD47')
    make_kpi_card('J', 3, 'L', '本月出库批次',
                  f'=SUMPRODUCT((出库记录!C2:C501<>"")*(出库记录!E2:E501>0)*(MONTH(出库记录!A2:A501)=MONTH(TODAY()))*(YEAR(出库记录!A2:A501)=YEAR(TODAY())))',
                  'ED7D31')

    # ---- KPI 卡片 (第2行: 7-9) ----
    make_kpi_card('B', 7, 'D', '本月入库总量',
                  f'=SUMPRODUCT((入库记录!E2:E501)*(MONTH(入库记录!A2:A501)=MONTH(TODAY()))*(YEAR(入库记录!A2:A501)=YEAR(TODAY())))',
                  '4472C4')
    make_kpi_card('F', 7, 'H', '本月出库总量',
                  f'=SUMPRODUCT((出库记录!E2:E501)*(MONTH(出库记录!A2:A501)=MONTH(TODAY()))*(YEAR(出库记录!A2:A501)=YEAR(TODAY())))',
                  '70AD47')
    make_kpi_card('J', 7, 'L', '低库存预警数',
                  f'=COUNTIF(库存汇总!I2:I{NUM_INV_ROWS + 1},"低于安全库存")+COUNTIF(库存汇总!I2:I{NUM_INV_ROWS + 1},"缺货")',
                  'FF0000')

    # ---- 低库存预警列表 ----
    alert_start = 11
    ws.merge_cells(f'B{alert_start}:L{alert_start}')
    section = ws.cell(row=alert_start, column=2, value='⚠️ 低库存预警列表（库存 ≤ 安全库存）')
    section.font = Font(name='微软雅黑', size=12, bold=True, color='C00000')
    section.alignment = LEFT_WRAP
    ws.row_dimensions[alert_start].height = 28

    alert_header_row = alert_start + 1
    alert_headers = ['物料编码', '物料名称', '规格型号', '单位', '安全库存',
                     '当前库存', '库存状态']
    alert_cols_map = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    for i, h in enumerate(alert_headers):
        col_letter = alert_cols_map[i]
        cell = ws.cell(row=alert_header_row, column=ord(col_letter) - ord('A') + 1, value=h)
        cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[alert_header_row].height = 24

    # 预警列表数据（最多显示 20 行）
    ALERT_ROWS = 20
    for i in range(ALERT_ROWS):
        r = alert_header_row + 1 + i
        row_ref = i + 2  # 库存汇总表的行号（从第2行开始）

        # 使用 IF+SMALL 提取非正常的物料
        # B列：物料编码
        ws.cell(row=r, column=2,
                value=f'=IFERROR(INDEX(库存汇总!A:A,SMALL(IF((库存汇总!$I$2:$I${NUM_INV_ROWS + 1}<>"正常")*(库存汇总!$A$2:$A${NUM_INV_ROWS + 1}<>""),ROW(库存汇总!$A$2:$A${NUM_INV_ROWS + 1})),{i + 1})),"")')
        # 名称
        ws.cell(row=r, column=3,
                value=f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},物料主数据!A:G,2,FALSE),""))')
        # 规格
        ws.cell(row=r, column=4,
                value=f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},物料主数据!A:G,3,FALSE),""))')
        # 单位
        ws.cell(row=r, column=5,
                value=f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},物料主数据!A:G,4,FALSE),""))')
        # 安全库存
        ws.cell(row=r, column=6,
                value=f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},物料主数据!A:G,6,FALSE),""))')
        # 当前库存
        ws.cell(row=r, column=7,
                value=f'=IF(B{r}="","",SUMIF(库存汇总!A:A,B{r},库存汇总!H:H))')
        # 库存状态
        ws.cell(row=r, column=8,
                value=f'=IF(B{r}="","",IF(G{r}<=0,"缺货",IF(G{r}<=F{r},"低于安全库存","正常")))')

        for c_offset in range(7):
            c = 2 + c_offset
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = CENTER
            if (r - alert_header_row) % 2 == 0:
                ws.cell(row=r, column=c).fill = ODD_FILL
            else:
                ws.cell(row=r, column=c).fill = EVEN_FILL

        ws.row_dimensions[r].height = 20

    # 预警条件格式
    alert_range = f'B{alert_header_row + 1}:H{alert_header_row + ALERT_ROWS}'
    ws.conditional_formatting.add(
        alert_range,
        FormulaRule(
            formula=[f'AND($B{alert_header_row + 1}<>"", $H{alert_header_row + 1}="缺货")'],
            fill=DANGER_FILL,
            font=Font(name='微软雅黑', size=10, color='9C0006', bold=True)
        )
    )
    ws.conditional_formatting.add(
        alert_range,
        FormulaRule(
            formula=[f'AND($B{alert_header_row + 1}<>"", $H{alert_header_row + 1}="低于安全库存")'],
            fill=WARN_FILL,
            font=Font(name='微软雅黑', size=10, color='9C6500', bold=True)
        )
    )

    # ---- 月度趋势表 ----
    trend_start = alert_header_row + ALERT_ROWS + 3
    ws.merge_cells(f'B{trend_start}:L{trend_start}')
    section = ws.cell(row=trend_start, column=2, value='📈 近6个月出入库趋势')
    section.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
    section.alignment = LEFT_WRAP
    ws.row_dimensions[trend_start].height = 28

    trend_headers = ['月份', '入库批次', '入库总量', '出库批次', '出库总量', '净增库存']
    trend_header_row = trend_start + 1
    for i, h in enumerate(trend_headers):
        col_letter = ['B', 'D', 'F', 'H', 'J', 'L'][i]
        cell = ws.cell(row=trend_header_row, column=ord(col_letter) - ord('A') + 1, value=h)
        cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 合并月份单元格（每个月份对应一整行）
    for i in range(6):
        r = trend_header_row + 1 + i
        # 月份标签（前推5个月到当前月）
        month_offset = 5 - i  # i=0 是5个月前, i=5 是当月
        ws.merge_cells(f'B{r}:C{r}')
        ws.cell(row=r, column=2,
                value=f'=TEXT(DATE(YEAR(TODAY()),MONTH(TODAY())-{month_offset},1),"YYYY年M月")')
        ws.cell(row=r, column=2).font = BOLD_FONT
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3).border = THIN_BORDER

        mm = month_offset
        # 入库批次
        ws.merge_cells(f'D{r}:E{r}')
        ws.cell(row=r, column=4,
                value=f'=SUMPRODUCT((入库记录!C2:C501<>"")*(入库记录!E2:E501>0)*(MONTH(入库记录!A2:A501)=MONTH(EDATE(TODAY(),-{mm})))*(YEAR(入库记录!A2:A501)=YEAR(EDATE(TODAY(),-{mm}))))')
        ws.cell(row=r, column=4).font = DATA_FONT
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=5).border = THIN_BORDER

        # 入库总量
        ws.merge_cells(f'F{r}:G{r}')
        ws.cell(row=r, column=6,
                value=f'=SUMPRODUCT((入库记录!E2:E501)*(MONTH(入库记录!A2:A501)=MONTH(EDATE(TODAY(),-{mm})))*(YEAR(入库记录!A2:A501)=YEAR(EDATE(TODAY(),-{mm}))))')
        ws.cell(row=r, column=6).font = DATA_FONT
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=6).border = THIN_BORDER
        ws.cell(row=r, column=7).border = THIN_BORDER

        # 出库批次
        ws.merge_cells(f'H{r}:I{r}')
        ws.cell(row=r, column=8,
                value=f'=SUMPRODUCT((出库记录!C2:C501<>"")*(出库记录!E2:E501>0)*(MONTH(出库记录!A2:A501)=MONTH(EDATE(TODAY(),-{mm})))*(YEAR(出库记录!A2:A501)=YEAR(EDATE(TODAY(),-{mm}))))')
        ws.cell(row=r, column=8).font = DATA_FONT
        ws.cell(row=r, column=8).alignment = CENTER
        ws.cell(row=r, column=8).border = THIN_BORDER
        ws.cell(row=r, column=9).border = THIN_BORDER

        # 出库总量
        ws.merge_cells(f'J{r}:K{r}')
        ws.cell(row=r, column=10,
                value=f'=SUMPRODUCT((出库记录!E2:E501)*(MONTH(出库记录!A2:A501)=MONTH(EDATE(TODAY(),-{mm})))*(YEAR(出库记录!A2:A501)=YEAR(EDATE(TODAY(),-{mm}))))')
        ws.cell(row=r, column=10).font = DATA_FONT
        ws.cell(row=r, column=10).alignment = CENTER
        ws.cell(row=r, column=10).border = THIN_BORDER
        ws.cell(row=r, column=11).border = THIN_BORDER

        # 净增库存 = 入库总量 - 出库总量
        ws.cell(row=r, column=12,
                value=f'=IF(AND(F{r}="",J{r}=""),"",(F{r}&"")-(J{r}&""))')
        # 简化：=F{r}-J{r}
        # 重新用更简单的方式
        ws.cell(row=r, column=12,
                value=f'=SUMPRODUCT((入库记录!E2:E501)*(MONTH(入库记录!A2:A501)=MONTH(EDATE(TODAY(),-{mm})))*(YEAR(入库记录!A2:A501)=YEAR(EDATE(TODAY(),-{mm}))))-SUMPRODUCT((出库记录!E2:E501)*(MONTH(出库记录!A2:A501)=MONTH(EDATE(TODAY(),-{mm})))*(YEAR(出库记录!A2:A501)=YEAR(EDATE(TODAY(),-{mm}))))')
        ws.cell(row=r, column=12).font = DATA_FONT
        ws.cell(row=r, column=12).alignment = CENTER
        ws.cell(row=r, column=12).border = THIN_BORDER

        # 行背景色
        fill = EVEN_FILL if i % 2 == 1 else ODD_FILL
        for c_idx in range(2, 13):
            if ws.cell(row=r, column=c_idx).fill == PatternFill():
                ws.cell(row=r, column=c_idx).fill = fill

        ws.row_dimensions[r].height = 22

    # ---- 页脚 ----
    footer_row = trend_header_row + 8
    ws.merge_cells(f'B{footer_row}:L{footer_row}')
    ws.cell(row=footer_row, column=2,
            value=f'数据更新时间：{date.today()}  |  所有数据由公式自动计算，无需手动刷新').font = Font(name='微软雅黑', size=9, italic=True, color='999999')
    ws.cell(row=footer_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # 隐藏网格线
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = 'FF0000'

    return ws


# ============================================================================
# 保护设置
# ============================================================================

def apply_protection(wb):
    """对所有工作表应用保护：锁定公式，解锁输入区域"""

    # --- 物料主数据：解锁 A-G 列（全部可编辑），隐藏 H 列 ---
    ws = wb['物料主数据']
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = True  # 允许插入行
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = True
    ws.protection.sort = True
    ws.protection.autoFilter = True
    for r in range(2, 202):
        for c in range(1, 8):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    # --- 入库记录：解锁输入列 (A,B,C,E,F,G,H,I)，锁定 D（公式列）---
    ws = wb['入库记录']
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = True
    ws.protection.deleteRows = True
    ws.protection.sort = True
    ws.protection.autoFilter = True
    for r in range(2, 502):
        for c in [1, 2, 3, 5, 6, 7, 8, 9]:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    # --- 出库记录：解锁输入列 (A,B,C,E,F,G,H)，锁定 D（公式列）---
    ws = wb['出库记录']
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = True
    ws.protection.deleteRows = True
    ws.protection.sort = True
    ws.protection.autoFilter = True
    for r in range(2, 502):
        for c in [1, 2, 3, 5, 6, 7, 8]:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    # --- 库存汇总：全部锁定（纯公式表）---
    ws = wb['库存汇总']
    ws.protection.sheet = True

    # --- 仪表盘：全部锁定 ---
    ws = wb['仪表盘']
    ws.protection.sheet = True

    # --- 操作说明：锁定 ---
    ws = wb['操作说明']
    ws.protection.sheet = True


# ============================================================================
# 主函数
# ============================================================================

def main():
    # 解析命令行参数
    clear_mode = '--clear' in sys.argv
    sample_data = not clear_mode

    print('=' * 60)
    print('  来料出入库管理系统 - Excel 生成工具')
    if clear_mode:
        print('  [清除模式] 生成空白模板，不含示例数据')
    print('=' * 60)

    wb = Workbook()

    # 按顺序构建各工作表
    print('[1/6] 构建「操作说明」...')
    build_instructions(wb)

    print('[2/6] 构建「物料主数据」...')
    build_material_master(wb, sample_data=sample_data)

    print('[3/6] 构建「入库记录」...')
    build_inbound_records(wb, sample_data=sample_data)

    print('[4/6] 构建「出库记录」...')
    build_outbound_records(wb, sample_data=sample_data)

    print('[5/6] 构建「库存汇总」（公式计算表）...')
    build_inventory_summary(wb)

    print('[6/6] 构建「仪表盘」...')
    build_dashboard(wb)

    # 应用保护
    print('应用工作表保护...')
    apply_protection(wb)

    # 设置活动工作表为操作说明
    wb.active = wb.sheetnames.index('操作说明')

    # 保存
    print(f'\n保存文件到: {OUTPUT_PATH}')
    wb.save(OUTPUT_PATH)

    print('\n' + '=' * 60)
    print('  [OK] 生成完成!')
    print(f'  文件位置: {OUTPUT_PATH}')
    print('  请使用 WPS Office 打开使用。')
    print('=' * 60)


if __name__ == '__main__':
    main()
